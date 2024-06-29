import pandas as pd
import numpy as np
import openpyxl
import math
#导入工作簿对象
workbook = openpyxl.load_workbook('C:\\Users\\30794\\OneDrive\\桌面\\score.xlsx')
sheetNames = workbook.sheetnames
for sheetName in sheetNames:
    sheet = workbook[sheetName]
    n = sheet.max_column
    if n == 9:
        sheet.cell(row=1,column=n+1).value = '总分'
        sheet.cell(row=1, column=n + 2).value = '平均分'
        sheet.cell(row=1, column=n + 3).value = '标准差'
        for i in range(2,sheet.max_row+1):
            sum = 0
            list1 = list()
            sum = 0
            for j in range(3,10):
                sum += sheet.cell(row=i,column=j).value
                list1.append(sheet.cell(row=i,column=j).value)
            sheet.cell(row=i,column=n+1).value = sum
            sheet.cell(row=i, column=n + 2).value = sum/7
            elm_o = 0
            for z in range(len(list1)):
                elm_o += abs(list1[z]-sum/7)**2
            sheet.cell(row=i, column=n + 3).value = (elm_o/7)**(1/2)
    else : break



workbook.save("C:\\Users\\30794\\OneDrive\\桌面\\score.xlsx")
